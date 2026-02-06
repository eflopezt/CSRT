"""
Utilidades para generar archivos Excel con validaciones y múltiples hojas.
"""
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
import logging

logger = logging.getLogger('personal')


def crear_excel_con_validaciones(datos_principales, nombre_hoja_principal, catalogos=None, columnas_validacion=None):
    """
    Crea un archivo Excel con datos principales y hojas de catálogos con validaciones.
    
    Args:
        datos_principales: DataFrame con los datos principales
        nombre_hoja_principal: Nombre de la hoja principal
        catalogos: Dict con {nombre_hoja: DataFrame} de catálogos
        columnas_validacion: Dict con {columna: nombre_hoja_catalogo} para validaciones
    
    Returns:
        BytesIO con el archivo Excel
    """
    output = BytesIO()
    
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Escribir hoja principal
            datos_principales.to_excel(writer, index=False, sheet_name=nombre_hoja_principal)
            
            # Escribir hojas de catálogos
            if catalogos:
                for nombre_catalogo, df_catalogo in catalogos.items():
                    df_catalogo.to_excel(writer, index=False, sheet_name=nombre_catalogo)
            
            # Aplicar estilos y validaciones
            workbook = writer.book
            hoja_principal = writer.sheets[nombre_hoja_principal]
            
            # Estilo del encabezado
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in hoja_principal[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar ancho de columnas y aplicar formato de texto a DNI
            columnas_texto = ['NroDoc', 'DNI', 'Responsable_DNI', 'CodigoFotocheck', 'Celular']
            
            for column in hoja_principal.columns:
                max_length = 0
                column_letter = column[0].column_letter
                column_name = column[0].value
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                    
                    # Aplicar formato de texto a columnas específicas (DNI, etc.)
                    if column_name in columnas_texto and cell.row > 1:  # Skip header
                        cell.number_format = FORMAT_TEXT
                        # Asegurar que el valor se mantiene como string
                        if cell.value is not None:
                            cell.value = str(cell.value)
                
                adjusted_width = min(max_length + 2, 50)
                hoja_principal.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar validaciones de datos
            if columnas_validacion and catalogos:
                for columna, catalogo_nombre in columnas_validacion.items():
                    # Buscar el índice de la columna
                    col_idx = None
                    for idx, cell in enumerate(hoja_principal[1], start=1):
                        if cell.value == columna:
                            col_idx = idx
                            break
                    
                    if col_idx and catalogo_nombre in catalogos:
                        # Obtener la hoja del catálogo
                        hoja_catalogo = writer.sheets[catalogo_nombre]
                        
                        # Contar filas del catálogo
                        num_filas_catalogo = catalogos[catalogo_nombre].shape[0]
                        
                        # Crear validación de datos (lista desplegable)
                        from openpyxl.utils import get_column_letter
                        col_letter = get_column_letter(col_idx)
                        
                        # Fórmula para la lista (referencia a la hoja de catálogo)
                        formula = f"'{catalogo_nombre}'!$A$2:$A${num_filas_catalogo + 1}"
                        
                        dv = DataValidation(
                            type="list",
                            formula1=formula,
                            allow_blank=True,
                            showErrorMessage=True,
                            errorTitle='Valor inválido',
                            error='Por favor, selecciona un valor de la lista'
                        )
                        
                        # Aplicar validación a las filas de datos (desde fila 2 hasta 1000)
                        dv.add(f'{col_letter}2:{col_letter}1000')
                        hoja_principal.add_data_validation(dv)
            
            # Congelar primera fila
            hoja_principal.freeze_panes = 'A2'
        
        logger.info(f"Excel creado exitosamente: {nombre_hoja_principal}")
    except Exception as e:
        logger.error(f"Error al crear Excel: {str(e)}")
        raise
    
    output.seek(0)
    return output


def crear_plantilla_personal(personal_queryset=None):
    """
    Crea una plantilla Excel para importar/actualizar Personal con catálogos y validaciones.
    """
    from .models import Area, SubArea, Personal
    
    # Datos principales (personal actual si se proporciona)
    if personal_queryset:
        data_personal = []
        for p in personal_queryset:
            data_personal.append({
                'NroDoc': p.nro_doc,
                'TipoDoc': p.tipo_doc,
                'ApellidosNombres': p.apellidos_nombres,
                'CodigoFotocheck': p.codigo_fotocheck,
                'Cargo': p.cargo,
                'TipoTrabajador': p.tipo_trab,
                'SubArea': p.subarea.nombre if p.subarea else '',
                'Estado': p.estado,
                'FechaAlta': p.fecha_alta.strftime('%Y-%m-%d') if p.fecha_alta else '',
                'FechaCese': p.fecha_cese.strftime('%Y-%m-%d') if p.fecha_cese else '',
                'FechaNacimiento': p.fecha_nacimiento.strftime('%Y-%m-%d') if p.fecha_nacimiento else '',
                'Sexo': p.sexo,
                'Celular': p.celular,
                'CorreoPersonal': p.correo_personal,
                'CorreoCorporativo': p.correo_corporativo,
                'Direccion': p.direccion,
                'Ubigeo': p.ubigeo,
                'RegimenLaboral': p.regimen_laboral,
                'RegimenTurno': p.regimen_turno,
                'DiasLibresCorte2025': float(p.dias_libres_corte_2025),
                'Observaciones': p.observaciones,
            })
    else:
        # Plantilla vacía con una fila de ejemplo
        data_personal = [{
            'NroDoc': '12345678',
            'TipoDoc': 'DNI',
            'ApellidosNombres': 'APELLIDOS, NOMBRES',
            'CodigoFotocheck': '',
            'Cargo': 'CARGO EJEMPLO',
            'TipoTrabajador': 'Empleado',
            'SubArea': '',
            'Estado': 'Activo',
            'FechaAlta': '2024-01-01',
            'FechaCese': '',
            'FechaNacimiento': '1990-01-01',
            'Sexo': 'M',
            'Celular': '',
            'CorreoPersonal': '',
            'CorreoCorporativo': '',
            'Direccion': '',
            'Ubigeo': '',
            'RegimenLaboral': '',
            'RegimenTurno': '21x7',
            'DiasLibresCorte2025': 0,
            'Observaciones': '',
        }]
    
    df_personal = pd.DataFrame(data_personal)
    
    # Catálogos
    catalogos = {
        'CAT_SubAreas': pd.DataFrame({
            'SubArea': [a.nombre for a in SubArea.objects.filter(activa=True).order_by('nombre')],
            'Area': [a.area.nombre for a in SubArea.objects.filter(activa=True).order_by('nombre')]
        }),
        'CAT_TipoDoc': pd.DataFrame({
            'TipoDoc': ['DNI', 'CE', 'Pasaporte']
        }),
        'CAT_TipoTrabajador': pd.DataFrame({
            'TipoTrabajador': ['Empleado', 'Obrero']
        }),
        'CAT_Estado': pd.DataFrame({
            'Estado': ['Activo', 'Inactivo', 'Suspendido', 'Cesado']
        }),
        'CAT_Sexo': pd.DataFrame({
            'Sexo': ['M', 'F']
        }),
    }
    
    # Columnas con validación
    columnas_validacion = {
        'SubArea': 'CAT_SubAreas',
        'TipoDoc': 'CAT_TipoDoc',
        'TipoTrabajador': 'CAT_TipoTrabajador',
        'Estado': 'CAT_Estado',
        'Sexo': 'CAT_Sexo',
    }
    
    return crear_excel_con_validaciones(df_personal, 'Personal', catalogos, columnas_validacion)


def crear_plantilla_gerencias(gerencias_queryset=None):
    """
    Crea una plantilla Excel para importar/actualizar Gerencias.
    """
    from .models import Area, Personal
    
    # Datos principales
    if gerencias_queryset:
        data = []
        for g in gerencias_queryset:
            responsables = list(g.responsables.all())
            data.append({
                'Nombre': g.nombre,
                'Responsable_DNI': ', '.join([p.nro_doc for p in responsables]),
                'Responsable_Nombre': ', '.join([p.apellidos_nombres for p in responsables]),
                'Descripcion': g.descripcion,
                'Activa': 'Sí' if g.activa else 'No',
            })
    else:
        data = [{
            'Nombre': 'AREA EJEMPLO',
            'Responsable_DNI': '',
            'Responsable_Nombre': '',
            'Descripcion': '',
            'Activa': 'Sí',
        }]
    
    df_gerencias = pd.DataFrame(data)
    
    # Catálogo de responsables
    catalogos = {
        'CAT_Responsables': pd.DataFrame({
            'DNI': [p.nro_doc for p in Personal.objects.filter(estado='Activo').order_by('apellidos_nombres')],
            'Nombre': [p.apellidos_nombres for p in Personal.objects.filter(estado='Activo').order_by('apellidos_nombres')]
        }),
        'CAT_Activa': pd.DataFrame({
            'Activa': ['Sí', 'No']
        }),
    }
    
    columnas_validacion = {
        'Responsable_DNI': 'CAT_Responsables',
        'Activa': 'CAT_Activa',
    }
    
    return crear_excel_con_validaciones(df_gerencias, 'Gerencias', catalogos, columnas_validacion)


def crear_plantilla_areas(areas_queryset=None):
    """
    Crea una plantilla Excel para importar/actualizar Áreas.
    """
    from .models import SubArea, Area
    
    # Datos principales
    if areas_queryset:
        data = []
        for a in areas_queryset:
            data.append({
                'Nombre': a.nombre,
                'Area': a.area.nombre,
                'Descripcion': a.descripcion,
                'Activa': 'Sí' if a.activa else 'No',
            })
    else:
        data = [{
            'Nombre': 'SUBAREA EJEMPLO',
            'Area': '',
            'Descripcion': '',
            'Activa': 'Sí',
        }]
    
    df_areas = pd.DataFrame(data)
    
    # Catálogos
    catalogos = {
        'CAT_Areas': pd.DataFrame({
            'Area': [g.nombre for g in Area.objects.filter(activa=True).order_by('nombre')]
        }),
        'CAT_Activa': pd.DataFrame({
            'Activa': ['Sí', 'No']
        }),
    }
    
    columnas_validacion = {
        'Area': 'CAT_Areas',
        'Activa': 'CAT_Activa',
    }
    
    return crear_excel_con_validaciones(df_areas, 'Areas', catalogos, columnas_validacion)


def crear_plantilla_roster(mes, anio, personal_queryset, rosters_queryset=None):
    """
    Crea una plantilla Excel para importar/actualizar Roster con validaciones.
    """
    from calendar import monthrange
    from datetime import datetime
    from collections import defaultdict
    
    dias_en_mes = monthrange(anio, mes)[1]
    
    # Organizar roster existente
    roster_dict = defaultdict(dict)
    if rosters_queryset:
        for r in rosters_queryset:
            roster_dict[r.personal_id][r.fecha.day] = r.codigo
    
    # Datos principales
    data = []
    for persona in personal_queryset:
        fila = {
            'DNI': persona.nro_doc,
            'ApellidosNombres': persona.apellidos_nombres,
            'SubArea': persona.subarea.nombre if persona.subarea else '',
            'DiasLibresCorte2025': float(persona.dias_libres_corte_2025),
        }
        
        # Agregar columnas de días
        for dia in range(1, dias_en_mes + 1):
            codigo = roster_dict[persona.id].get(dia, '')
            fila[f'Dia{dia:02d}'] = codigo
        
        data.append(fila)
    
    df_roster = pd.DataFrame(data)
    
    # Catálogos
    catalogos = {
        'CAT_Codigos': pd.DataFrame({
            'Codigo': ['T', 'TR', 'DL', 'DOL', 'DM', 'V', 'F', 'FC', 'P', 'I', 'L']
        }),
        'CAT_Descripcion': pd.DataFrame({
            'Codigo': ['T', 'TR', 'DL', 'DOL', 'DM', 'V', 'F', 'FC', 'P', 'I', 'L'],
            'Descripcion': [
                'Trabajo Presencial',
                'Trabajo Remoto',
                'Día Libre',
                'Compensación por Horario Extendido',
                'Descanso Médico',
                'Vacaciones',
                'Feriado No Recuperable',
                'Feriado Compensable',
                'Permiso',
                'Inasistencia',
                'Licencia'
            ]
        }),
    }
    
    # Aplicar validación a todas las columnas de días
    columnas_validacion = {}
    for dia in range(1, dias_en_mes + 1):
        columnas_validacion[f'Dia{dia:02d}'] = 'CAT_Codigos'
    
    return crear_excel_con_validaciones(df_roster, 'Roster', catalogos, columnas_validacion)
